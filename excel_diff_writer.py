"""
Excel 하이라이트 모듈
- BomDiff 결과를 Excel 워크시트에 시각적으로 표시
- 변경사항 요약 시트 생성
"""
from copy import copy
from typing import Dict, List, Optional, Tuple

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from models import BomRow
from utils import clean_text, normalize_header
from excel_template import find_bom_header_row_and_cols
from bom_comparator import BomDiff, RowDiff


# 하이라이트 색상: 달라진 것 = 노랑, 없어진 것 = 빨강
FILL_HIGHLIGHT = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_REMOVED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FONT_STRIKETHROUGH = Font(strikethrough=True, color="999999")
FONT_REMOVED_HEADER = Font(bold=True, color="9C0006", strikethrough=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _find_row_in_sheet(
    ws: Worksheet,
    row: BomRow,
    start_row: int,
    col_product: int,
    col_material: int,
    col_usage: int,
) -> Optional[int]:
    """워크시트에서 BomRow에 해당하는 Excel 행 번호를 찾기"""
    prod = (row.product or "").strip()
    mat = (row.material_name or "").strip()
    usage = (row.usage or "").strip()

    for r in range(start_row, ws.max_row + 1):
        cell_prod = clean_text(ws.cell(r, col_product).value)
        cell_mat = clean_text(ws.cell(r, col_material).value)
        cell_usage = clean_text(ws.cell(r, col_usage).value)
        if cell_prod == prod and cell_mat == mat and cell_usage == usage:
            return r

    for r in range(start_row, ws.max_row + 1):
        cell_prod = clean_text(ws.cell(r, col_product).value)
        cell_mat = clean_text(ws.cell(r, col_material).value)
        if cell_prod == prod and cell_mat == mat:
            return r

    return None


def apply_highlights(
    ws: Worksheet,
    diff: BomDiff,
    color_headers: List[str],
) -> int:
    """
    워크시트에 변경사항 하이라이트 적용.
    Returns: 삭제된 행이 추가된 개수
    """
    header_row, col_map = find_bom_header_row_and_cols(ws)
    start_row = header_row + 1

    c_category = col_map.get("category")
    c_product = col_map["product"]
    c_material = col_map["materialname"]
    c_usage = col_map["usage"]
    c_quality = col_map["qualitydetails"]
    c_supplier = col_map.get("supplierallocate") or col_map.get("supplier")
    c_supp_art = col_map["supplierarticlenumber"]
    if c_supplier is None:
        c_supplier = c_quality

    c_color_start = c_supplier + 1

    color_col_map: Dict[str, int] = {}
    for j, h in enumerate(color_headers):
        color_col_map[h] = c_color_start + j

    max_data_col = max(
        c_supplier,
        c_color_start + len(color_headers) - 1 if color_headers else c_supplier,
    )

    field_to_col = {
        "Category": c_category,
        "Product": c_product,
        "Material Name": c_material,
        "Supplier Article Number": c_supp_art,
        "Usage": c_usage,
        "Quality Details": c_quality,
        "Supplier": c_supplier,
    }

    # --- 1) 추가된 행 하이라이트 (초록) ---
    for added_row in diff.added_rows:
        excel_r = _find_row_in_sheet(ws, added_row, start_row, c_product, c_material, c_usage)
        if excel_r is None:
            continue
        for c in range(c_category or c_product, max_data_col + 1):
            ws.cell(excel_r, c).fill = FILL_HIGHLIGHT

    # --- 2) 변경된 행의 변경 셀 하이라이트 (노랑) ---
    for rd in diff.modified_rows:
        excel_r = _find_row_in_sheet(ws, rd.new_row, start_row, c_product, c_material, c_usage)
        if excel_r is None:
            continue

        for field_name, (old_val, new_val) in rd.changed_fields.items():
            col = field_to_col.get(field_name)
            if col:
                ws.cell(excel_r, col).fill = FILL_HIGHLIGHT

        for color_header, (old_val, new_val) in rd.changed_colors.items():
            col = color_col_map.get(color_header)
            if col:
                ws.cell(excel_r, col).fill = FILL_HIGHLIGHT

    # --- 3) 삭제된 행 (하단에 취소선으로 추가) ---
    removed_count = 0
    if diff.removed_rows:
        last_data_row = start_row
        for r in range(start_row, ws.max_row + 1):
            if clean_text(ws.cell(r, c_product).value):
                last_data_row = r

        insert_row = last_data_row + 2

        label_col = c_category if c_category else c_product
        label_cell = ws.cell(insert_row, label_col)
        label_cell.value = "[ 삭제된 자재 ]"
        label_cell.font = Font(bold=True, color="9C0006")
        insert_row += 1

        for removed_row in diff.removed_rows:
            r = insert_row
            if c_category:
                ws.cell(r, c_category).value = removed_row.category
            ws.cell(r, c_product).value = removed_row.product
            ws.cell(r, c_material).value = removed_row.material_name
            ws.cell(r, c_supp_art).value = removed_row.supplier_article_number
            ws.cell(r, c_usage).value = removed_row.usage
            ws.cell(r, c_quality).value = removed_row.quality_details
            ws.cell(r, c_supplier).value = removed_row.supplier

            for h, v in (removed_row.colors or {}).items():
                col = color_col_map.get(h)
                if col and v:
                    ws.cell(r, col).value = v

            for c in range(c_category or c_product, max_data_col + 1):
                cell = ws.cell(r, c)
                cell.fill = FILL_REMOVED
                cell.font = FONT_STRIKETHROUGH
                cell.border = THIN_BORDER

            insert_row += 1
            removed_count += 1

    # --- 4) 컬러 헤더 하이라이트 ---
    for h in diff.added_colors:
        col = color_col_map.get(h)
        if col:
            ws.cell(header_row, col).fill = FILL_HIGHLIGHT

    return removed_count


def create_summary_sheet(wb, diff: BomDiff, sheet_name: str = "변경사항 요약"):
    """변경사항 요약 시트 생성"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    row = 1
    ws.cell(row, 1).value = "BOM 변경사항 요약"
    ws.cell(row, 1).font = Font(bold=True, size=14)
    row += 2

    if not diff.has_changes:
        ws.cell(row, 1).value = "변경사항 없음"
        return

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 35

    # --- Master 변경 ---
    if diff.master_changes:
        ws.cell(row, 1).value = "Master 정보 변경"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        row += 1
        ws.cell(row, 1).value = "필드"
        ws.cell(row, 2).value = "이전 값"
        ws.cell(row, 3).value = "새 값"
        for c in range(1, 4):
            ws.cell(row, c).font = Font(bold=True)
            ws.cell(row, c).border = THIN_BORDER
        row += 1
        for k, (ov, nv) in diff.master_changes.items():
            ws.cell(row, 1).value = k
            ws.cell(row, 2).value = ov
            ws.cell(row, 3).value = nv
            ws.cell(row, 2).fill = FILL_REMOVED
            ws.cell(row, 3).fill = FILL_HIGHLIGHT
            for c in range(1, 4):
                ws.cell(row, c).border = THIN_BORDER
            row += 1
        row += 1

    # --- 컬러 변경 ---
    if diff.added_colors or diff.removed_colors:
        ws.cell(row, 1).value = "컬러 변경"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        row += 1
        for h in diff.added_colors:
            ws.cell(row, 1).value = "추가"
            ws.cell(row, 1).fill = FILL_HIGHLIGHT
            ws.cell(row, 2).value = h.replace("\n", " / ")
            ws.cell(row, 2).fill = FILL_HIGHLIGHT
            row += 1
        for h in diff.removed_colors:
            ws.cell(row, 1).value = "삭제"
            ws.cell(row, 1).fill = FILL_REMOVED
            ws.cell(row, 2).value = h.replace("\n", " / ")
            ws.cell(row, 2).fill = FILL_REMOVED
            ws.cell(row, 2).font = FONT_STRIKETHROUGH
            row += 1
        row += 1

    # --- 추가된 자재 ---
    if diff.added_rows:
        ws.cell(row, 1).value = f"추가된 자재 ({len(diff.added_rows)}행)"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        row += 1
        for lbl, ci in [("Section", 1), ("Product", 2), ("Material Name", 3), ("Usage", 4)]:
            ws.cell(row, ci).value = lbl
            ws.cell(row, ci).font = Font(bold=True)
            ws.cell(row, ci).border = THIN_BORDER
        row += 1
        for r in diff.added_rows:
            ws.cell(row, 1).value = r.category
            ws.cell(row, 2).value = r.product
            ws.cell(row, 3).value = r.material_name
            ws.cell(row, 4).value = r.usage
            for c in range(1, 5):
                ws.cell(row, c).fill = FILL_HIGHLIGHT
                ws.cell(row, c).border = THIN_BORDER
            row += 1
        row += 1

    # --- 삭제된 자재 ---
    if diff.removed_rows:
        ws.cell(row, 1).value = f"삭제된 자재 ({len(diff.removed_rows)}행)"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        row += 1
        for lbl, ci in [("Section", 1), ("Product", 2), ("Material Name", 3), ("Usage", 4)]:
            ws.cell(row, ci).value = lbl
            ws.cell(row, ci).font = Font(bold=True)
            ws.cell(row, ci).border = THIN_BORDER
        row += 1
        for r in diff.removed_rows:
            ws.cell(row, 1).value = r.category
            ws.cell(row, 2).value = r.product
            ws.cell(row, 3).value = r.material_name
            ws.cell(row, 4).value = r.usage
            for c in range(1, 5):
                ws.cell(row, c).fill = FILL_REMOVED
                ws.cell(row, c).font = FONT_STRIKETHROUGH
                ws.cell(row, c).border = THIN_BORDER
            row += 1
        row += 1

    # --- 변경된 행 상세 ---
    if diff.modified_rows:
        ws.cell(row, 1).value = f"변경된 행 ({len(diff.modified_rows)}행)"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        row += 1
        for lbl, ci in [("Product", 1), ("Material", 2), ("변경 필드", 3), ("이전 값", 4), ("새 값", 5)]:
            ws.cell(row, ci).value = lbl
            ws.cell(row, ci).font = Font(bold=True)
            ws.cell(row, ci).border = THIN_BORDER
        row += 1

        for rd in diff.modified_rows:
            first = True
            all_changes = list(rd.changed_fields.items()) + [
                (h.replace("\n", " / "), vals) for h, vals in rd.changed_colors.items()
            ]
            for field_name, (ov, nv) in all_changes:
                if first:
                    ws.cell(row, 1).value = rd.new_row.product
                    ws.cell(row, 2).value = rd.new_row.material_name
                    first = False
                ws.cell(row, 3).value = field_name
                ws.cell(row, 4).value = ov
                ws.cell(row, 4).fill = FILL_REMOVED
                ws.cell(row, 5).value = nv
                ws.cell(row, 5).fill = FILL_HIGHLIGHT
                for c in range(1, 6):
                    ws.cell(row, c).border = THIN_BORDER
                row += 1
            row += 1
