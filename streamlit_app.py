"""
BOM PDF → Excel Auto Filler + 비교 분석  (Streamlit Web App)

모드:
1. BOM 분석 자동화  — PDF → Excel 변환 (단일/복수)
2. Updated BOM 비교 — 이전 Excel + 새 PDF → 변경사항 하이라이트
"""

import os
import sys
import tempfile

import streamlit as st
from openpyxl import load_workbook

# ── [테스트] IP 감지 확인 ─────────────────────────────────────
# 이 코드는 IP가 정상 감지되는지 확인 후 제거 예정
try:
    headers = st.context.headers
    client_ip = headers.get("X-Forwarded-For", "감지불가").split(",")[0].strip()
    st.sidebar.caption(f"감지된 IP: {client_ip}")
except Exception as e:
    st.sidebar.caption(f"IP 감지 실패: {e}")
# ── [테스트] 끝 ───────────────────────────────────────────────

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
if _APP_DIR not in sys.path:
    sys.path.insert(0, _APP_DIR)

from excel_writer import fill_sheet, sanitize_sheet_name
from pdf_parser import parse_master_from_pdf, extract_bom_rows_from_pdf
from models import group_rows_by_material
from excel_reader import read_excel_bom
from bom_comparator import compare_boms
from excel_diff_writer import apply_highlights, create_summary_sheet

# ── 페이지 설정 ──────────────────────────────────────────────
st.set_page_config(
    page_title="BOM PDF → Excel 자동 입력",
    page_icon="📋",
    layout="centered",
)

DEFAULT_TEMPLATE = os.path.join(_APP_DIR, "양식.xlsx")

# ── Session State 초기화 ─────────────────────────────────────
if "result_mode1" not in st.session_state:
    st.session_state.result_mode1 = None
if "logs_mode1" not in st.session_state:
    st.session_state.logs_mode1 = []
if "result_mode2" not in st.session_state:
    st.session_state.result_mode2 = None
if "logs_mode2" not in st.session_state:
    st.session_state.logs_mode2 = []

# ── 사이드바: 모드 선택 ──────────────────────────────────────
with st.sidebar:
    st.header("모드 선택")
    mode = st.radio(
        "",
        ["📋 BOM 분석 자동화", "🔄 Updated BOM 비교"],
        label_visibility="collapsed",
    )

    st.divider()

    if mode == "📋 BOM 분석 자동화":
        st.markdown(
            "**BOM PDF를 Excel 양식에 자동 입력**\n\n"
            "- 단일/복수 PDF 지원\n"
            "- 이미지, 컬러값 자동 매핑\n"
            "- 복수 PDF → 시트별 분리"
        )
    else:
        st.markdown(
            "**이전 Excel과 새 PDF를 비교하여**\n"
            "**변경사항을 자동 감지**\n\n"
            "- 🟡 변경/추가된 셀 → 노랑\n"
            "- 🔴 삭제된 행 → 빨강 취소선\n"
            "- 변경사항 요약 시트 자동 생성"
        )

    st.divider()
    st.caption("v2.0")

# ══════════════════════════════════════════════════════════════
# 모드 1: BOM 분석 자동화
# ══════════════════════════════════════════════════════════════
if mode == "📋 BOM 분석 자동화":
    st.title("📋 BOM 분석 자동화")

    # ── 1) Excel 양식 선택 ──
    st.subheader("1. Excel 양식")
    has_default = os.path.exists(DEFAULT_TEMPLATE)
    template_options = (
        ["기본 내장 양식 (양식.xlsx)", "직접 업로드"]
        if has_default else ["직접 업로드"]
    )
    template_option = st.radio(
        "양식 선택", options=template_options, horizontal=True, key="tpl_opt_1",
        label_visibility="collapsed",
    )

    uploaded_template = None
    if template_option == "직접 업로드":
        uploaded_template = st.file_uploader("Excel 양식 (.xlsx)", type=["xlsx"], key="tpl_1")

    # ── 2) BOM PDF 업로드 ──
    st.subheader("2. BOM PDF")
    uploaded_pdfs = st.file_uploader(
        "PDF 파일 (복수 선택 가능)",
        type=["pdf"],
        accept_multiple_files=True,
        key="pdfs_1",
    )

    # ── 3) 실행 ──
    can_run = bool(uploaded_pdfs)
    if template_option == "직접 업로드" and uploaded_template is None:
        can_run = False

    st.divider()

    if st.button("실행", disabled=not can_run, use_container_width=True, type="primary", key="run_1"):
        st.session_state.result_mode1 = None
        st.session_state.logs_mode1 = []
        logs = st.session_state.logs_mode1
        total = len(uploaded_pdfs)

        with st.status(f"{total}개 PDF 변환 중...", expanded=True) as status:
            progress = st.progress(0, text="준비 중...")

            tmpdir = tempfile.mkdtemp()
            try:
                if template_option == "직접 업로드":
                    tpl_path = os.path.join(tmpdir, "template.xlsx")
                    with open(tpl_path, "wb") as f:
                        f.write(uploaded_template.getvalue())
                else:
                    tpl_path = DEFAULT_TEMPLATE

                pdf_paths = []
                for i, pdf_file in enumerate(uploaded_pdfs):
                    p = os.path.join(tmpdir, f"{i}.pdf")
                    with open(p, "wb") as f:
                        f.write(pdf_file.getvalue())
                    pdf_paths.append(p)

                if total == 1:
                    pdf_name = uploaded_pdfs[0].name
                    base_name = os.path.splitext(pdf_name)[0]
                    out_name = f"{base_name}_filled.xlsx"
                    out_path = os.path.join(tmpdir, out_name)

                    progress.progress(0, text=f"처리 중: {pdf_name}")
                    logs.append(f"[1/1] {pdf_name}")

                    try:
                        wb = load_workbook(tpl_path)
                        ws = wb.active
                        fill_sheet(ws, pdf_paths[0])
                        wb.save(out_path)
                        wb.close()
                        logs.append(f"   완료: {out_name}")
                        with open(out_path, "rb") as f:
                            st.session_state.result_mode1 = (out_name, f.read())
                    except Exception as e:
                        logs.append(f"   실패: {e}")
                        st.error(f"실패: {e}")

                    progress.progress(1.0, text="완료!")
                else:
                    wb = load_workbook(tpl_path)
                    original_sheets = list(wb.sheetnames)
                    template_ws = wb.active
                    sheet_names_used = set()
                    success_count = 0
                    fail_count = 0

                    for idx, (pdf_path, pdf_file) in enumerate(zip(pdf_paths, uploaded_pdfs)):
                        pdf_name = pdf_file.name
                        progress.progress(idx / total, text=f"[{idx + 1}/{total}] {pdf_name}")
                        logs.append(f"[{idx + 1}/{total}] {pdf_name}")
                        try:
                            new_ws = wb.copy_worksheet(template_ws)
                            design_number = fill_sheet(new_ws, pdf_path)
                            name = design_number or os.path.splitext(pdf_name)[0]
                            name = sanitize_sheet_name(name)
                            base_name = name
                            counter = 1
                            while name in sheet_names_used:
                                s = f"_{counter}"
                                name = sanitize_sheet_name(base_name[:31 - len(s)] + s)
                                counter += 1
                            sheet_names_used.add(name)
                            new_ws.title = name
                            logs.append(f"   완료 → 시트: {name}")
                            success_count += 1
                        except Exception as e:
                            logs.append(f"   실패: {e}")
                            fail_count += 1

                    for sn in original_sheets:
                        if sn in wb.sheetnames:
                            wb.remove(wb[sn])

                    out_name = "BOM_combined_filled.xlsx"
                    out_path = os.path.join(tmpdir, out_name)
                    wb.save(out_path)
                    wb.close()
                    with open(out_path, "rb") as f:
                        st.session_state.result_mode1 = (out_name, f.read())
                    progress.progress(1.0, text="완료!")
                    logs.append(f"성공 {success_count}개 / 실패 {fail_count}개")

            finally:
                import shutil
                shutil.rmtree(tmpdir, ignore_errors=True)

            status.update(label="완료!", state="complete")

    # ── 결과 다운로드 (실행 완료 후에만) ──
    if st.session_state.result_mode1:
        fname, fbytes = st.session_state.result_mode1
        st.success(f"✅ {fname}")
        st.download_button(
            label="📥 다운로드",
            data=fbytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    if st.session_state.logs_mode1:
        with st.expander("처리 로그", expanded=False):
            st.code("\n".join(st.session_state.logs_mode1), language=None)

# ══════════════════════════════════════════════════════════════
# 모드 2: Updated BOM 비교
# ══════════════════════════════════════════════════════════════
else:
    st.title("🔄 Updated BOM 비교")

    # ── 1) Excel 양식 선택 ──
    st.subheader("1. Excel 양식")
    has_default = os.path.exists(DEFAULT_TEMPLATE)
    template_options = (
        ["기본 내장 양식 (양식.xlsx)", "직접 업로드"]
        if has_default else ["직접 업로드"]
    )
    template_option = st.radio(
        "양식 선택", options=template_options, horizontal=True, key="tpl_opt_2",
        label_visibility="collapsed",
    )

    uploaded_template = None
    if template_option == "직접 업로드":
        uploaded_template = st.file_uploader("Excel 양식 (.xlsx)", type=["xlsx"], key="tpl_2")

    # ── 2) 이전 Excel ──
    st.subheader("2. 이전 Excel")
    uploaded_prev_excel = st.file_uploader(
        "이전에 생성한 BOM Excel",
        type=["xlsx"],
        key="prev_excel_2",
    )

    # ── 3) 새 BOM PDF ──
    st.subheader("3. 새 BOM PDF")
    uploaded_new_pdf = st.file_uploader(
        "업데이트된 BOM PDF",
        type=["pdf"],
        accept_multiple_files=False,
        key="new_pdf_2",
    )

    # ── 4) 실행 ──
    can_run = bool(uploaded_prev_excel and uploaded_new_pdf)
    if template_option == "직접 업로드" and uploaded_template is None:
        can_run = False

    st.divider()

    if st.button("비교 실행", disabled=not can_run, use_container_width=True, type="primary", key="run_2"):
        st.session_state.result_mode2 = None
        st.session_state.logs_mode2 = []
        logs = st.session_state.logs_mode2

        with st.status("비교 분석 중...", expanded=True) as status:
            progress = st.progress(0, text="준비 중...")

            tmpdir = tempfile.mkdtemp()
            try:
                if template_option == "직접 업로드":
                    tpl_path = os.path.join(tmpdir, "template.xlsx")
                    with open(tpl_path, "wb") as f:
                        f.write(uploaded_template.getvalue())
                else:
                    tpl_path = DEFAULT_TEMPLATE

                prev_path = os.path.join(tmpdir, "prev.xlsx")
                with open(prev_path, "wb") as f:
                    f.write(uploaded_prev_excel.getvalue())

                pdf_path = os.path.join(tmpdir, "new.pdf")
                with open(pdf_path, "wb") as f:
                    f.write(uploaded_new_pdf.getvalue())

                base_name = os.path.splitext(uploaded_new_pdf.name)[0]
                out_name = f"{base_name}_compared.xlsx"
                out_path = os.path.join(tmpdir, out_name)

                try:
                    progress.progress(0.2, text="PDF 변환 중...")
                    logs.append(f"새 PDF: {uploaded_new_pdf.name}")

                    wb = load_workbook(tpl_path)
                    ws = wb.active
                    fill_sheet(ws, pdf_path)
                    logs.append("   BOM 데이터 입력 완료")

                    progress.progress(0.4, text="이전 Excel 읽는 중...")
                    old_rows, old_colors, old_master = read_excel_bom(prev_path)
                    logs.append(f"   이전: {len(old_rows)}행, {len(old_colors)}컬러")

                    progress.progress(0.6, text="비교 중...")
                    new_master = parse_master_from_pdf(pdf_path)
                    raw_rows, new_colors = extract_bom_rows_from_pdf(pdf_path)
                    new_rows = group_rows_by_material(raw_rows)
                    logs.append(f"   새로: {len(new_rows)}행, {len(new_colors)}컬러")

                    diff = compare_boms(
                        old_rows, old_colors, old_master,
                        new_rows, new_colors, new_master,
                    )

                    progress.progress(0.8, text="하이라이트 적용 중...")
                    if diff.has_changes:
                        apply_highlights(ws, diff, new_colors)
                        create_summary_sheet(wb, diff)
                        for line in diff.summary_lines():
                            logs.append(line)
                        logs.append("   하이라이트 적용 완료")
                    else:
                        logs.append("   변경사항 없음")

                    wb.save(out_path)
                    wb.close()
                    with open(out_path, "rb") as f:
                        st.session_state.result_mode2 = (out_name, f.read())

                except Exception as e:
                    logs.append(f"   실패: {e}")
                    st.error(f"실패: {e}")

                progress.progress(1.0, text="완료!")

            finally:
                import shutil
                shutil.rmtree(tmpdir, ignore_errors=True)

            status.update(label="비교 완료!", state="complete")

    # ── 결과 다운로드 (실행 완료 후에만) ──
    if st.session_state.result_mode2:
        fname, fbytes = st.session_state.result_mode2
        st.success(f"✅ {fname}")
        st.download_button(
            label="📥 다운로드",
            data=fbytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    if st.session_state.logs_mode2:
        with st.expander("처리 로그", expanded=False):
            st.code("\n".join(st.session_state.logs_mode2), language=None)
