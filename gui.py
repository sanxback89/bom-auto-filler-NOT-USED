"""
GUI 모듈 - tkinter 기반 사용자 인터페이스
+ BOM 비교 분석 기능 통합
"""
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook

from excel_writer import fill_sheet, sanitize_sheet_name
from pdf_parser import parse_master_from_pdf, extract_bom_rows_from_pdf
from models import group_rows_by_material

from excel_reader import read_excel_bom
from bom_comparator import compare_boms
from excel_diff_writer import apply_highlights, create_summary_sheet


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("BOM PDF → Excel Template Auto Filler (Multi) + 비교 분석")
        self.geometry("700x580")

        self.template_path = tk.StringVar()
        self.prev_excel_path = tk.StringVar()
        self.saved_template = None
        self.saved_prev_excel = None

        self._build_ui()

    def _build_ui(self):
        pad = 8

        # Template
        tk.Label(self, text="1) 엑셀 양식(.xlsx) - 한 번만 선택하면 재사용됩니다").grid(
            row=0, column=0, sticky="w", padx=pad, pady=(pad, 2)
        )
        tk.Entry(self, textvariable=self.template_path, width=65).grid(
            row=1, column=0, padx=pad, sticky="w"
        )
        tk.Button(self, text="찾기", command=self.browse_template, width=10).grid(
            row=1, column=1, padx=pad
        )

        # PDF (복수 선택)
        tk.Label(self, text="2) BOM PDF(.pdf) - 여러 파일 선택 가능, 선택 시 자동 실행됩니다").grid(
            row=2, column=0, sticky="w", padx=pad, pady=(pad, 2)
        )
        tk.Entry(self, text="", width=65, state="readonly").grid(
            row=3, column=0, padx=pad, sticky="w"
        )
        tk.Button(self, text="찾기", command=self.browse_pdfs_and_run, width=10).grid(
            row=3, column=1, padx=pad
        )

        # 구분선
        ttk.Separator(self, orient="horizontal").grid(
            row=4, column=0, columnspan=2, sticky="ew", padx=pad, pady=(pad, 2)
        )

        # 이전 Excel (비교용) — 선택사항
        tk.Label(
            self,
            text="3) [선택사항] 비교할 이전 Excel(.xlsx) — 선택 시 변경사항 하이라이트",
            fg="#0066CC",
        ).grid(row=5, column=0, sticky="w", padx=pad, pady=(2, 2))

        compare_frame = tk.Frame(self)
        compare_frame.grid(row=6, column=0, columnspan=2, padx=pad, sticky="ew")
        tk.Entry(compare_frame, textvariable=self.prev_excel_path, width=55).pack(side=tk.LEFT)
        tk.Button(compare_frame, text="찾기", command=self.browse_prev_excel, width=10).pack(side=tk.LEFT, padx=(6, 0))
        tk.Button(compare_frame, text="초기화", command=self.clear_prev_excel, width=8).pack(side=tk.LEFT, padx=(4, 0))

        # Progress bar
        progress_frame = tk.Frame(self)
        progress_frame.grid(row=7, column=0, columnspan=2, padx=pad, pady=(pad, 2), sticky="ew")

        self.progress = ttk.Progressbar(progress_frame, mode="determinate", length=550)
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.progress_label = tk.Label(progress_frame, text="", width=12, anchor="e")
        self.progress_label.pack(side=tk.RIGHT, padx=(6, 0))

        # Log
        tk.Label(self, text="로그").grid(row=8, column=0, sticky="w", padx=pad, pady=(pad, 2))

        scroll_frame = tk.Frame(self)
        scroll_frame.grid(row=9, column=0, columnspan=2, padx=pad, pady=(2, pad), sticky="nsew")

        scrollbar = tk.Scrollbar(scroll_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.log = tk.Text(scroll_frame, height=15, width=80, yscrollcommand=scrollbar.set)
        self.log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log.yview)

        # 태그 설정 (컬러 로그)
        self.log.tag_configure("added", foreground="#006100")
        self.log.tag_configure("removed", foreground="#9C0006")
        self.log.tag_configure("changed", foreground="#9C6500")
        self.log.tag_configure("header", foreground="#0066CC", font=("TkDefaultFont", 10, "bold"))

        self.grid_rowconfigure(9, weight=1)
        self.grid_columnconfigure(0, weight=1)

    def _reset_progress(self):
        self.progress["value"] = 0
        self.progress_label.config(text="")
        self.update_idletasks()

    def _set_progress(self, current: int, total: int):
        pct = int(current / total * 100) if total else 0
        self.progress["value"] = pct
        self.progress_label.config(text=f"{current}/{total}")
        self.update_idletasks()

    def browse_template(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.template_path.set(path)
            self.saved_template = path
            self._log(f"✅ 템플릿 선택됨: {os.path.basename(path)}")
            self._log("   → 이 템플릿은 앞으로 계속 사용됩니다\n")

    def browse_prev_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.prev_excel_path.set(path)
            self.saved_prev_excel = path
            self._log(f"🔄 비교 대상 Excel 선택됨: {os.path.basename(path)}")
            self._log("   → PDF 분석 후 이 파일과 비교하여 변경사항을 표시합니다\n")

    def clear_prev_excel(self):
        self.prev_excel_path.set("")
        self.saved_prev_excel = None
        self._log("   비교 모드 해제됨 (일반 변환 모드)\n")

    def browse_pdfs_and_run(self):
        """복수 PDF 선택 후 자동으로 실행 - 하나의 파일, 시트별 분리"""
        if not self.saved_template or not os.path.exists(self.saved_template):
            self._log("⚠️  먼저 엑셀 양식을 선택해주세요!\n")
            messagebox.showwarning("템플릿 필요", "먼저 엑셀 양식 파일을 선택해주세요.")
            return

        paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if not paths:
            return

        compare_mode = bool(self.saved_prev_excel and os.path.exists(self.saved_prev_excel))
        total = len(paths)
        self._reset_progress()

        self._log("=" * 70)
        self._log(f"📋 선택된 PDF: {total}개" + (" [🔄 비교 모드]" if compare_mode else ""))
        for i, p in enumerate(paths, 1):
            self._log(f"   {i}. {os.path.basename(p)}")
        self._log("=" * 70 + "\n")

        output_dir = os.path.dirname(paths[0])

        try:
            if total == 1:
                self._process_single_pdf(paths[0], output_dir, compare_mode)
            else:
                self._process_multi_pdfs(paths, output_dir, compare_mode)
        except Exception as e:
            self._log(f"   ❌ 실패: {str(e)}")
            self._log("=" * 70 + "\n")
            messagebox.showerror("오류", f"처리 중 오류 발생:\n\n{str(e)}")

    def _process_single_pdf(self, pdf_path: str, output_dir: str, compare_mode: bool):
        """
        단일 PDF 처리.
        워크북을 한 번만 열고 fill_sheet → 하이라이트 → save.
        이미지가 메모리에 유지된 상태로 저장되므로 손실 없음.
        """
        pdf_basename = os.path.splitext(os.path.basename(pdf_path))[0]
        suffix = "_compared" if compare_mode else "_filled"
        output_path = os.path.join(output_dir, f"{pdf_basename}{suffix}.xlsx")

        self._log(f"📄 [1/1] 처리 중: {os.path.basename(pdf_path)}")
        self._set_progress(0, 1)

        # ── 1) 워크북 열기 + fill_sheet (이미지 포함, 기존 로직 그대로) ──
        wb = load_workbook(self.saved_template)
        ws = wb.active
        fill_sheet(ws, pdf_path)
        self._set_progress(1, 1)
        self._log(f"   ✅ BOM 데이터 입력 완료")

        # ── 2) 비교 모드: 같은 워크북에 하이라이트 적용 ──
        if compare_mode:
            diff, new_colors = self._run_comparison(pdf_path)
            if diff and diff.has_changes:
                self._log("\n   📝 Excel에 하이라이트 적용 중...")
                apply_highlights(ws, diff, new_colors)
                create_summary_sheet(wb, diff)
                self._log("   ✅ 하이라이트 적용 완료!")
                self._log("   📊 '변경사항 요약' 시트가 추가되었습니다.")

        # ── 3) 한 번만 저장 ──
        wb.save(output_path)

        self._log("\n" + "=" * 70)
        self._log(f"📊 작업 완료!")
        self._log(f"   📁 저장 위치: {output_path}")
        self._log("=" * 70 + "\n")
        messagebox.showinfo("완료", f"작업 완료!\n\n저장: {os.path.basename(output_path)}")

    def _process_multi_pdfs(self, paths, output_dir: str, compare_mode: bool):
        """복수 PDF → 하나의 파일, 시트별 분리 (기존 로직 그대로)"""
        total = len(paths)
        output_path = os.path.join(output_dir, "BOM_combined_filled.xlsx")

        wb = load_workbook(self.saved_template)
        original_sheet_names = list(wb.sheetnames)
        template_ws = wb.active

        sheet_names_used = set()
        success_count = 0
        fail_count = 0

        for idx, pdf_path in enumerate(paths, 1):
            self._set_progress(idx - 1, total)
            self._log(f"📄 [{idx}/{total}] 처리 중: {os.path.basename(pdf_path)}")

            try:
                new_ws = wb.copy_worksheet(template_ws)
                design_number = fill_sheet(new_ws, pdf_path)

                # 시트 이름 결정
                name = design_number or os.path.splitext(os.path.basename(pdf_path))[0]
                name = sanitize_sheet_name(name)

                base_name = name
                counter = 1
                while name in sheet_names_used:
                    s = f"_{counter}"
                    name = sanitize_sheet_name(base_name[:31 - len(s)] + s)
                    counter += 1
                sheet_names_used.add(name)
                new_ws.title = name

                self._log(f"   ✅ 완료 → 시트: {name}")
                success_count += 1

            except Exception as e:
                self._log(f"   ❌ 실패: {str(e)}")
                fail_count += 1

            self._set_progress(idx, total)

        # 원본 템플릿 시트 모두 삭제
        for sn in original_sheet_names:
            if sn in wb.sheetnames:
                wb.remove(wb[sn])

        wb.save(output_path)

        if fail_count > 0:
            self._log(f"\n   ⚠️ 성공: {success_count}개 / 실패: {fail_count}개")

        if compare_mode:
            self._log("\n   ℹ️ 비교 분석은 단일 PDF 모드에서만 지원됩니다.")

        self._log("\n" + "=" * 70)
        self._log(f"📊 작업 완료!")
        self._log(f"   📁 저장 위치: {output_path}")
        self._log("=" * 70 + "\n")
        messagebox.showinfo("완료", f"작업 완료!\n\n저장: {os.path.basename(output_path)}")

    def _run_comparison(self, new_pdf_path: str):
        """
        비교 분석 실행. diff + color_headers만 반환.
        하이라이트 적용은 호출자가 같은 워크북에서 처리.

        Returns:
            (BomDiff, new_color_headers) or (None, [])
        """
        self._log("\n" + "-" * 50)
        self._log("🔄 비교 분석 시작...", tag="header")

        try:
            # 1) 이전 Excel 읽기
            self._log("   이전 Excel 읽는 중...")
            old_rows, old_colors, old_master = read_excel_bom(self.saved_prev_excel)
            self._log(f"   이전: {len(old_rows)}행, {len(old_colors)}개 컬러")

            # 2) 새 PDF 파싱 (BomRow 필요)
            self._log("   새 BOM 파싱 중...")
            new_master = parse_master_from_pdf(new_pdf_path)
            raw_rows, new_colors = extract_bom_rows_from_pdf(new_pdf_path)
            new_rows = group_rows_by_material(raw_rows)
            self._log(f"   새로: {len(new_rows)}행, {len(new_colors)}개 컬러")

            # 3) 비교
            self._log("   비교 중...")
            diff = compare_boms(
                old_rows, old_colors, old_master,
                new_rows, new_colors, new_master,
            )

            # 4) 결과 로그 출력
            if not diff.has_changes:
                self._log("\n   ✅ 변경사항 없음!", tag="header")
            else:
                self._log("\n   📋 [변경사항 요약]", tag="header")
                for line in diff.summary_lines():
                    if line.strip().startswith("+") or "추가" in line:
                        self._log(line, tag="added")
                    elif line.strip().startswith("-") or "삭제" in line:
                        self._log(line, tag="removed")
                    elif line.strip().startswith("~") or "변경" in line or "\u2192" in line:
                        self._log(line, tag="changed")
                    else:
                        self._log(line)

            self._log("-" * 50)
            return diff, new_colors

        except Exception as e:
            self._log(f"   ❌ 비교 중 오류: {str(e)}")
            import traceback
            self._log(traceback.format_exc())
            return None, []

    def _log(self, msg: str, tag: str = None):
        if tag:
            self.log.insert("end", msg + "\n", tag)
        else:
            self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.update_idletasks()
