"""
기존 Excel 파일에서 BOM 데이터를 읽어오는 모듈
- fill_template으로 생성된 Excel을 다시 파싱하여 BomRow 리스트로 변환
- Master 필드, 컬러 헤더도 함께 추출
"""
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import BomRow
from utils import clean_text, normalize_header
from excel_template import (
    MASTER_LABELS,
    find_master_value_cells,
    find_bom_header_row_and_cols,
)


def read_master_from_excel(ws: Worksheet) -> Dict[str, str]:
    """Excel 워크시트에서 Master 필드값 읽기"""
    master = {}
    try:
        cells = find_master_value_cells(ws)
        for key, (r, c) in cells.items():
            val = clean_text(ws.cell(r, c).value)
            master[key] = val
    except ValueError:
        pass
    return master


def read_bom_rows_from_excel(ws: Worksheet) -> Tuple[List[BomRow], List[str]]:
    """
    Excel 워크시트에서 BOM 행과 컬러 헤더를 읽어 BomRow 리스트로 변환.

    Returns:
        (rows, color_headers)
    """
    header_row, col_map = find_bom_header_row_and_cols(ws)
    start_row = header_row + 1

    c_category = col_map.get("category")
    c_product = col_map["product"]
    c_material = col_map["materialname"]
    c_supp_art = col_map["supplierarticlenumber"]
    c_usage = col_map["usage"]
    c_quality = col_map["qualitydetails"]
    c_supplier = col_map.get("supplierallocate") or col_map.get("supplier")
    c_image = col_map.get("image")

    if c_supplier is None:
        c_supplier = c_quality

    c_color_start = c_supplier + 1

    # 컬러 헤더 읽기: c_color_start부터 빈 셀 또는 시트 끝까지
    color_headers: List[str] = []
    for c in range(c_color_start, ws.max_column + 1):
        header_val = clean_text(ws.cell(header_row, c).value)
        if not header_val:
            next_val = clean_text(ws.cell(header_row, c + 1).value) if c + 1 <= ws.max_column else ""
            if not next_val:
                break
            continue
        raw_val = ws.cell(header_row, c).value
        if raw_val is not None:
            color_headers.append(str(raw_val).strip())

    # 데이터 행 읽기
    rows: List[BomRow] = []
    empty_streak = 0

    for r in range(start_row, ws.max_row + 1):
        product = clean_text(ws.cell(r, c_product).value)
        material = clean_text(ws.cell(r, c_material).value)

        if not product and not material:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0

        # 서브타이틀 행 스킵 ([ Packaging and Labels ] 등)
        if product and product.startswith("[") and product.endswith("]"):
            continue
        category_val = clean_text(ws.cell(r, c_category).value) if c_category else ""
        if category_val and category_val.startswith("[") and category_val.endswith("]"):
            continue

        category = clean_text(ws.cell(r, c_category).value) if c_category else ""
        supp_art = clean_text(ws.cell(r, c_supp_art).value)
        usage = clean_text(ws.cell(r, c_usage).value)
        quality = clean_text(ws.cell(r, c_quality).value)
        supplier = clean_text(ws.cell(r, c_supplier).value)

        colors: Dict[str, str] = {}
        for j, h in enumerate(color_headers):
            c = c_color_start + j
            if c <= ws.max_column:
                val = clean_text(ws.cell(r, c).value)
                if val:
                    colors[h] = val

        rows.append(BomRow(
            category=category,
            product=product,
            material_name=material,
            supplier_article_number=supp_art,
            usage=usage,
            quality_details=quality,
            supplier=supplier,
            colors=colors,
            image_png=None,
        ))

    return rows, color_headers


def read_excel_bom(excel_path: str, sheet_name: Optional[str] = None) -> Tuple[
    List[BomRow], List[str], Dict[str, str]
]:
    """
    Excel 파일에서 BOM 데이터 전체를 읽기.

    Returns:
        (rows, color_headers, master)
    """
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    master = read_master_from_excel(ws)
    rows, color_headers = read_bom_rows_from_excel(ws)
    wb.close()

    return rows, color_headers, master
